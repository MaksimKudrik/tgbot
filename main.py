import sqlite3
import asyncio
import logging
import os
from io import StringIO
from dotenv import load_dotenv
import pandas as pd
from aiogram import Bot, Dispatcher, Router
from aiogram.filters import Command, CommandStart
from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import CallbackQuery
from openpyxl import load_workbook
import csv

# Настройка логирования
logging.basicConfig(
    filename='bot.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Загрузка токена из .env
load_dotenv()
API_TOKEN = os.getenv("API_TOKEN")
if not API_TOKEN:
    logger.error("API_TOKEN не найден в .env файле")
    raise ValueError("API_TOKEN не задан в .env файле")

# Инициализация бота
bot = Bot(token=API_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot=bot, storage=storage)
router = Router()
dp.include_router(router)

# Путь к базе данных SQLite
DB_PATH = 'gym_bot.db'


# Инициализация базы данных
def init_db():
    """Инициализирует базу данных SQLite и создаёт таблицу users, если она не существует."""
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    user_id INTEGER PRIMARY KEY,
                    bench_press REAL DEFAULT 0.0,
                    squat REAL DEFAULT 0.0,
                    deadlift REAL DEFAULT 0.0
                )
            ''')
            conn.commit()
            logger.info("База данных инициализирована")
    except sqlite3.Error as e:
        logger.error(f"Ошибка при инициализации базы данных: {e}")
        raise


# Функции для работы с базой данных
def save_user_data(user_id: int, data: dict) -> None:
    """Сохраняет данные пользователя в базу данных."""
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT OR REPLACE INTO users (user_id, bench_press, squat, deadlift)
                VALUES (?, ?, ?, ?)
            ''', (
                user_id,
                data.get('bench_press', 0.0),
                data.get('squat', 0.0),
                data.get('deadlift', 0.0)
            ))
            conn.commit()
            logger.info(f"Данные пользователя {user_id} сохранены: {data}")
    except sqlite3.Error as e:
        logger.error(f"Ошибка при сохранении данных пользователя {user_id}: {e}")
        raise


def load_user_data(user_id: int) -> dict:
    """Загружает данные пользователя из базы данных."""
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT bench_press, squat, deadlift FROM users WHERE user_id = ?', (user_id,))
            result = cursor.fetchone()
            if result:
                return {
                    'bench_press': result[0],
                    'squat': result[1],
                    'deadlift': result[2]
                }
            return {'bench_press': 0.0, 'squat': 0.0, 'deadlift': 0.0}
    except sqlite3.Error as e:
        logger.error(f"Ошибка при загрузке данных пользователя {user_id}: {e}")
        return {'bench_press': 0.0, 'squat': 0.0, 'deadlift': 0.0}


def clear_user_data(user_id: int) -> None:
    """Очищает данные пользователя в базе данных."""
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM users WHERE user_id = ?', (user_id,))
            conn.commit()
            logger.info(f"Данные пользователя {user_id} очищены")
    except sqlite3.Error as e:
        logger.error(f"Ошибка при очистке данных пользователя {user_id}: {e}")
        raise


# Создание reply-клавиатуры для команд
def get_reply_command_keyboard() -> ReplyKeyboardMarkup:
    """Создаёт reply-клавиатуру с кнопками для команд."""
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="неделя")],
            [KeyboardButton(text="результаты"), KeyboardButton(text="сбросить")],
            [KeyboardButton(text="помощь")]
        ],
        resize_keyboard=True,
        is_persistent=True
    )
    return keyboard


# Создание reply-клавиатуры с днями и кнопкой "Назад"
def get_days_only_keyboard() -> ReplyKeyboardMarkup:
    """Создаёт reply-клавиатуру с кнопками для дней и кнопкой 'Назад'."""
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="понедельник"), KeyboardButton(text="среда"), KeyboardButton(text="пятница")],
            [KeyboardButton(text="Назад")]
        ],
        resize_keyboard=True,
        is_persistent=True
    )
    return keyboard


# Создание inline-клавиатуры для выбора недели
def get_week_keyboard() -> InlineKeyboardMarkup:
    """Создаёт inline-клавиатуру с кнопками для выбора недели."""
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=str(i), callback_data=f"week_{i}") for i in range(1, 5)],
        [InlineKeyboardButton(text=str(i), callback_data=f"week_{i}") for i in range(5, 9)]
    ])
    return keyboard


# Определение машины состояний для ввода максимальных весов и выбора недели
class MaxLiftForm(StatesGroup):
    bench_press = State()
    squat = State()
    deadlift = State()
    week_selection = State()


def csv_from_excel():
    """Конвертирует Excel-файл 'Муж высокий 3дневный.xlsx' в CSV с правильной структурой."""
    excel_file = 'training.xlsx'
    if not os.path.exists(excel_file):
        logger.error(f"Файл {excel_file} не найден")
        raise FileNotFoundError(f"Файл {excel_file} не найден")

    # Загружаем Excel-файл
    wb = load_workbook(excel_file)
    sheet = wb.active

    # Подготовка данных для CSV
    data = []
    current_week = None
    current_day = None

    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not any(row):  # Пропускаем пустые строки
            continue
        if isinstance(row[0], str) and row[0].startswith('неделя'):
            current_week = int(row[0].split()[1])
            continue
        if row[0] in ['понедельник', 'среда', 'пятница']:
            current_day = row[0]
            continue
        if row[0] and current_week and current_day:
            exercise = row[0]
            intensity = row[1] if row[1] else 'средняя'  # Значение по умолчанию для интенсивности
            reps = row[2] if row[2] else '3х8-12'  # Значение по умолчанию для подходов
            # Исправляем опечатку
            exercise = 'передняя дельта' if exercise == 'прередняя дельта' else exercise
            data.append([current_week, current_day, exercise, intensity, reps])

    # Записываем данные в CSV
    with open('training.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['неделя', 'день', 'упражнения', 'интенсивность', 'подходы х повторения'])
        writer.writerows(data)

    logger.info(f"Файл {excel_file} успешно конвертирован в training.csv")


# Загрузка данных из CSV
try:
    csv_from_excel()
    df = pd.read_csv('training.csv', encoding='utf-8')
    logger.info("CSV данные успешно загружены")
except Exception as e:
    logger.error(f"Ошибка при чтении CSV данных: {e}")
    raise

# Маппинг упражнений для расчёта весов
EXERCISE_MAPPING = {
    "жим лёжа": {"main_lift": "bench_press", "scale": 1.0, "min_weight": 20.0, "max_weight": 500.0, "increment": 2.5},
    "присед со штангой": {"main_lift": "squat", "scale": 1.0, "min_weight": 20.0, "max_weight": 600.0,
                          "increment": 2.5},
    "классическая тяга": {"main_lift": "deadlift", "scale": 1.0, "min_weight": 20.0, "max_weight": 700.0,
                          "increment": 2.5},
    "тяга вертикального блока": {"main_lift": "bench_press", "scale": 0.55, "min_weight": 10.0, "max_weight": 120.0,
                                 "increment": 2.5},
    "тяга горизонтального блока": {"main_lift": "bench_press", "scale": 0.55, "min_weight": 10.0, "max_weight": 120.0,
                                   "increment": 2.5},
    "шраги с гантелями": {"main_lift": "deadlift", "scale": 0.20, "min_weight": 4.0, "max_weight": 80.0,
                          "increment": 2.0},
    "косичка": {"main_lift": "bench_press", "scale": 0.50, "min_weight": 5.0, "max_weight": 50.0, "increment": 2.5},
    "французский жим лёжа": {"main_lift": "bench_press", "scale": 0.25, "min_weight": 10.0, "max_weight": 60.0,
                             "increment": 2.5},
    "сгибания на бицепс ez грифа": {"main_lift": "bench_press", "scale": 0.35, "min_weight": 5.0, "max_weight": 50.0,
                                    "increment": 2.5},
    "подъем на бицепс с прямым грифом": {"main_lift": "bench_press", "scale": 0.3, "min_weight": 10.0,
                                         "max_weight": 50.0, "increment": 2.5},
    "молотки": {"main_lift": "bench_press", "scale": 0.20, "min_weight": 4.0, "max_weight": 30.0, "increment": 2.0},
    "передняя дельта": {"main_lift": "bench_press", "scale": 0.18, "min_weight": 4.0, "max_weight": 25.0,
                        "increment": 2.0},
    "махи на среднюю дельту": {"main_lift": "bench_press", "scale": 0.15, "min_weight": 2.0, "max_weight": 25.0,
                               "increment": 2.0},
    "отведения на дельты": {"main_lift": "bench_press", "scale": 0.20, "min_weight": 10.0, "max_weight": 35.0,
                            "increment": 2.0},
    "подъем на носки в смите": {"main_lift": "squat", "scale": 0.3, "min_weight": 30.0, "max_weight": 100.0,
                                "increment": 5.0},
    "сгибание на предплечье": {"main_lift": "bench_press", "scale": 0.6, "min_weight": 25.0, "max_weight": 60.0,
                               "increment": 2.5},
    "разгибание на предплечье": {"main_lift": "bench_press", "scale": 0.6, "min_weight": 25.0, "max_weight": 60.0,
                                 "increment": 2.5},
    "разгибания ног в тренажере": {"main_lift": "squat", "scale": 0.60, "min_weight": 30.0, "max_weight": 100.0,
                                   "increment": 5.0},
    "сгибания ног в тренажере": {"main_lift": "squat", "scale": 0.60, "min_weight": 30.0, "max_weight": 90.0,
                                 "increment": 5.0},
    "жим гантелей лёжа 30°": {"main_lift": "bench_press", "scale": 0.30, "min_weight": 10.0, "max_weight": 50.0,
                              "increment": 2.0},
    "жим штанги": {"main_lift": "bench_press", "scale": 1.0, "min_weight": 20.0, "max_weight": 500.0, "increment": 2.5}
}


# Функция для расчёта веса
def calculate_weight(max_lift: float, intensity: str, exercise: str) -> str:
    """Рассчитывает вес на основе максимума, интенсивности и упражнения с фиксированными процентами."""
    try:
        mapping = EXERCISE_MAPPING.get(exercise.lower(), {
            "main_lift": "bench_press",
            "scale": 0.3,
            "min_weight": 5.0,
            "max_weight": 80.0,
            "increment": 2.5
        })
        scale = mapping["scale"]
        min_weight = mapping["min_weight"]
        max_weight = mapping["max_weight"]
        increment = mapping["increment"]

        base_weight = max_lift * scale
        intensity = intensity.lower()
        if intensity == "легкая":
            weight = base_weight * 0.60
        elif intensity == "средняя":
            weight = base_weight * 0.70
        elif intensity == "тяжёлая" or intensity == "тяжелая":  # Учитываем возможные варианты написания
            weight = base_weight * 0.80
        else:
            return "Не указан вес (неизвестная интенсивность)"

        weight = max(min_weight, min(max_weight, round(weight / increment) * increment))
        return f"{weight:.1f} кг"
    except Exception as e:
        logger.error(f"Ошибка при расчёте веса для упражнения {exercise}: {e}")
        return "Ошибка расчёта веса"


# Функция для форматирования программы тренировок
async def format_workout_plan(user_id: int, week: int, day: str) -> str:
    """Форматирует план тренировок для указанной недели и дня."""
    user_data = load_user_data(user_id)
    logger.info(f"Извлечены данные для пользователя {user_id}: {user_data}")

    max_weights = {
        "bench_press": user_data.get("bench_press", 0.0),
        "squat": user_data.get("squat", 0.0),
        "deadlift": user_data.get("deadlift", 0.0)
    }
    logger.info(f"Максимальные веса для пользователя {user_id}: {max_weights}")

    result = f"**Программа тренировок на {day.capitalize()}, Неделя {week}**\n\n"
    week_day_df = df[(df['неделя'] == week) & (df['день'] == day)]

    if week_day_df.empty:
        return f"Ошибка: Данные для недели {week}, дня {day} не найдены."

    for _, row in week_day_df.iterrows():
        exercise = row['упражнения']
        intensity = row['интенсивность']
        reps = row['подходы х повторения']

        mapping = EXERCISE_MAPPING.get(exercise.lower(), {"main_lift": "bench_press", "scale": 0.3})
        main_lift = mapping["main_lift"]
        max_lift = max_weights.get(main_lift, 0.0)

        weight = calculate_weight(max_lift, intensity,
                                  exercise) if max_lift > 0 else "Введите максимальные веса (/сбросить)"
        result += f"- {exercise}: {intensity.capitalize()} ({reps}, {weight})\n"

    return result


# Валидация ввода веса
def validate_weight(text: str) -> tuple[bool, float, str]:
    """Проверяет, является ли введённый текст допустимым весом."""
    try:
        weight = float(text)
        if weight < 0:
            return False, 0.0, "Вес не может быть отрицательным."
        if weight > 1000:
            return False, 0.0, "Вес слишком большой. Введите реалистичное значение (до 1000 кг)."
        return True, weight, ""
    except ValueError:
        return False, 0.0, "Введите число (например, 100) или 'пропустить'."


# Обработчик команды /start
@router.message(CommandStart())
async def start_command(message: Message, state: FSMContext):
    """Запускает процесс ввода максимальных весов."""
    try:
        await state.clear()
        await message.answer("Привет! Я бот для тренировок. Давай начнём с твоих максимальных результатов.")
        await message.answer("Введи свой максимальный вес в жиме лёжа (в кг, например, 100) или напиши 'пропустить':")
        await state.set_state(MaxLiftForm.bench_press)
        logger.info(f"Пользователь {message.from_user.id} начал ввод максимальных весов")
    except Exception as e:
        logger.error(f"Ошибка в start_command для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте снова с /start.")


# Обработчик команды /cancel
@router.message(Command("cancel"))
async def cancel_command(message: Message, state: FSMContext):
    """Отменяет текущий процесс ввода весов."""
    try:
        current_state = await state.get_state()
        if current_state is None:
            await message.answer("Нет активного процесса ввода весов.", reply_markup=get_reply_command_keyboard())
            logger.info(f"Пользователь {message.from_user.id} попытался отменить, но не было активного состояния")
            return
        await state.clear()
        await message.answer("Ввод весов отменён. Выбери действие:", reply_markup=get_reply_command_keyboard())
        logger.info(f"Пользователь {message.from_user.id} отменил ввод весов")
    except Exception as e:
        logger.error(f"Ошибка в cancel_command для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка при отмене. Попробуйте снова.",
                             reply_markup=get_reply_command_keyboard())


# Обработчик ввода максимума в жиме лёжа
@router.message(MaxLiftForm.bench_press)
async def process_bench_press(message: Message, state: FSMContext):
    """Обрабатывает ввод максимального веса в жиме лёжа."""
    try:
        if message.text.lower() == 'пропустить':
            await state.update_data(bench_press=0.0, squat=0.0, deadlift=0.0)
            user_data = await state.get_data()
            save_user_data(message.from_user.id, user_data)
            await state.clear()
            await message.answer(
                "Вы пропустили ввод весов. План будет без расчёта весов.\nВыбери действие:",
                reply_markup=get_reply_command_keyboard()
            )
            logger.info(f"Пользователь {message.from_user.id} пропустил ввод весов")
            return

        is_valid, weight, error_msg = validate_weight(message.text)
        if not is_valid:
            await message.answer(f"Ошибка: {error_msg} Для отмены используй /cancel.")
            return

        await state.update_data(bench_press=weight)
        await message.answer("Отлично! Теперь введи свой максимальный вес в приседе (в кг) или 'пропустить':")
        await state.set_state(MaxLiftForm.squat)
        logger.info(f"Пользователь {message.from_user.id} ввёл жим лёжа: {weight} кг")
    except Exception as e:
        logger.error(f"Ошибка в process_bench_press для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте снова или используй /cancel.")


# Обработчик ввода максимума в приседе
@router.message(MaxLiftForm.squat)
async def process_squat(message: Message, state: FSMContext):
    """Обрабатывает ввод максимального веса в приседе."""
    try:
        if message.text.lower() == 'пропустить':
            await state.update_data(squat=0.0, deadlift=0.0)
            user_data = await state.get_data()
            save_user_data(message.from_user.id, user_data)
            await state.clear()
            await message.answer(
                "Вы пропустили ввод весов. План будет без расчёта весов.\nВыбери действие:",
                reply_markup=get_reply_command_keyboard()
            )
            logger.info(f"Пользователь {message.from_user.id} пропустил ввод весов")
            return

        is_valid, weight, error_msg = validate_weight(message.text)
        if not is_valid:
            await message.answer(f"Ошибка: {error_msg} Для отмены используй /cancel.")
            return

        await state.update_data(squat=weight)
        await message.answer("Супер! Введи свой максимальный вес в становой тяге (в кг) или 'пропустить':")
        await state.set_state(MaxLiftForm.deadlift)
        logger.info(f"Пользователь {message.from_user.id} ввёл присед: {weight} кг")
    except Exception as e:
        logger.error(f"Ошибка в process_squat для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте снова или используй /cancel.")


# Обработчик ввода максимума в становой тяге
@router.message(MaxLiftForm.deadlift)
async def process_deadlift(message: Message, state: FSMContext):
    """Обрабатывает ввод максимального веса в становой тяге."""
    try:
        if message.text.lower() == 'пропустить':
            await state.update_data(deadlift=0.0)
            user_data = await state.get_data()
            save_user_data(message.from_user.id, user_data)
            await state.clear()
            await message.answer(
                f"Ввод завершён. Текущие веса:\n"
                f"Жим лёжа: {user_data.get('bench_press', 0.0)} кг\n"
                f"Присед: {user_data.get('squat', 0.0)} кг\n"
                f"Становая тяга: {user_data.get('deadlift', 0.0)} кг\n"
                "Выбери действие:",
                reply_markup=get_reply_command_keyboard()
            )
            logger.info(f"Пользователь {message.from_user.id} пропустил ввод становой тяги. Текущие веса: {user_data}")
            return

        is_valid, weight, error_msg = validate_weight(message.text)
        if not is_valid:
            await message.answer(f"Ошибка: {error_msg} Для отмены используй /cancel.")
            return

        await state.update_data(deadlift=weight)
        user_data = await state.get_data()
        save_user_data(message.from_user.id, user_data)
        await state.clear()
        await message.answer(
            f"Отлично, данные сохранены!\n"
            f"Жим лёжа: {user_data['bench_press']} кг\n"
            f"Присед: {user_data['squat']} кг\n"
            f"Становая тяга: {user_data['deadlift']} кг\n"
            "Выбери действие:",
            reply_markup=get_reply_command_keyboard()
        )
        logger.info(f"Пользователь {message.from_user.id} завершил ввод весов: {user_data}")
    except Exception as e:
        logger.error(f"Ошибка в process_deadlift для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте снова или используй /cancel.")


# Обработчик команды /результаты
@router.message(Command("результаты"))
async def my_weights_command(message: Message, state: FSMContext):
    """Отображает текущие максимальные веса пользователя."""
    try:
        user_data = load_user_data(message.from_user.id)
        logger.info(f"Проверка весов для пользователя {message.from_user.id}: {user_data}")
        if not user_data or all(value == 0.0 for value in user_data.values()):
            await message.answer(
                "Вы ещё не ввели максимальные веса. Используй /сбросить для ввода.",
                reply_markup=get_reply_command_keyboard()
            )
            logger.info(f"Пользователь {message.from_user.id} запросил веса, но данные отсутствуют")
            return
        await message.answer(
            f"Текущие максимальные веса:\n"
            f"Жим лёжа: {user_data.get('bench_press', 0.0)} кг\n"
            f"Присед: {user_data.get('squat', 0.0)} кг\n"
            f"Становая тяга: {user_data.get('deadlift', 0.0)} кг\n"
            "Выбери действие:",
            reply_markup=get_reply_command_keyboard()
        )
        logger.info(f"Пользователь {message.from_user.id} запросил текущие веса: {user_data}")
    except Exception as e:
        logger.error(f"Ошибка в my_weights_command для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка при получении весов. Попробуйте снова.",
                             reply_markup=get_reply_command_keyboard())


# Обработчик кнопки "результаты"
@router.message(lambda message: message.text == "результаты")
async def my_weights_button(message: Message, state: FSMContext):
    """Обрабатывает нажатие кнопки 'результаты'."""
    await my_weights_command(message, state)


# Обработчик команды /сбросить
@router.message(Command("сбросить"))
async def reset_command(message: Message, state: FSMContext):
    """Сбрасывает максимальные веса и начинает ввод заново."""
    try:
        await state.clear()
        clear_user_data(message.from_user.id)
        await message.answer(
            "Максимальные веса сброшены. Введи свой максимальный вес в жиме лёжа (в кг) или 'пропустить':")
        await state.set_state(MaxLiftForm.bench_press)
        logger.info(f"Пользователь {message.from_user.id} сбросил максимальные веса")
    except Exception as e:
        logger.error(f"Ошибка в reset_command для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка при сбросе весов. Попробуйте снова.",
                             reply_markup=get_reply_command_keyboard())


# Обработчик кнопки "сбросить"
@router.message(lambda message: message.text == "сбросить")
async def reset_button(message: Message, state: FSMContext):
    """Обрабатывает нажатие кнопки 'сбросить'."""
    await reset_command(message, state)


# Обработчик команды /неделя
@router.message(Command("неделя"))
async def week_command(message: Message, state: FSMContext):
    """Отображает inline-клавиатуру с выбором недели."""
    try:
        await state.clear()
        await message.answer(
            "Выбери неделю для тренировки:",
            reply_markup=get_week_keyboard()
        )
        logger.info(f"Пользователь {message.from_user.id} запросил выбор недели")
    except Exception as e:
        logger.error(f"Ошибка в week_command для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте снова.", reply_markup=get_reply_command_keyboard())


# Обработчик кнопки "неделя"
@router.message(lambda message: message.text == "неделя")
async def week_button(message: Message, state: FSMContext):
    """Обрабатывает нажатие кнопки 'неделя'."""
    await week_command(message, state)


# Обработчик callback для выбора недели
@router.callback_query(lambda c: c.data.startswith("week_"))
async def process_week_callback(callback: CallbackQuery, state: FSMContext):
    """Обрабатывает выбор недели и запрашивает день."""
    try:
        week = int(callback.data.split("_")[1])
        await state.update_data(selected_week=week)
        await state.set_state(MaxLiftForm.week_selection)
        await callback.message.edit_text(
            f"Выбрана неделя {week}. Выбери день для тренировки:",
            reply_markup=None
        )
        await callback.message.answer(
            "Выбери день:",
            reply_markup=get_days_only_keyboard()
        )
        logger.info(f"Пользователь {callback.from_user.id} выбрал неделю {week}")
    except Exception as e:
        logger.error(f"Ошибка в process_week_callback для пользователя {callback.from_user.id}: {e}")
        await callback.message.answer("Произошла ошибка. Попробуйте снова.", reply_markup=get_reply_command_keyboard())
    finally:
        await callback.answer()


# Обработчик команды /помощь
@router.message(Command("помощь"))
async def help_command(message: Message):
    """Отображает список доступных команд."""
    try:
        await message.answer(
            "Выбери действие из доступных команд:",
            reply_markup=get_reply_command_keyboard()
        )
        logger.info(f"Пользователь {message.from_user.id} запросил помощь")
    except Exception as e:
        logger.error(f"Ошибка в help_command для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка при отображении помощи. Попробуйте снова.",
                             reply_markup=get_reply_command_keyboard())


# Обработчик кнопки "помощь"
@router.message(lambda message: message.text == "помощь")
async def help_button(message: Message):
    """Обрабатывает нажатие кнопки 'помощь'."""
    await help_command(message)


# Обработчик кнопки "понедельник"
@router.message(MaxLiftForm.week_selection, lambda message: message.text == "понедельник")
async def monday_button(message: Message, state: FSMContext):
    """Обрабатывает нажатие кнопки 'понедельник'."""
    try:
        data = await state.get_data()
        week = data.get("selected_week")
        if not week:
            await message.answer("Ошибка: неделя не выбрана. Используй кнопку 'неделя'.",
                                 reply_markup=get_reply_command_keyboard())
            return
        workout_plan = await format_workout_plan(user_id=message.from_user.id, week=week, day="понедельник")
        await message.answer(
            workout_plan,
            reply_markup=get_days_only_keyboard()
        )
        logger.info(f"Пользователь {message.from_user.id} запросил программу на понедельник, неделя {week}")
    except Exception as e:
        logger.error(f"Ошибка в monday_button для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте снова.", reply_markup=get_reply_command_keyboard())


# Обработчик кнопки "среда"
@router.message(MaxLiftForm.week_selection, lambda message: message.text == "среда")
async def wednesday_button(message: Message, state: FSMContext):
    """Обрабатывает нажатие кнопки 'среда'."""
    try:
        data = await state.get_data()
        week = data.get("selected_week")
        if not week:
            await message.answer("Ошибка: неделя не выбрана. Используй кнопку 'неделя'.",
                                 reply_markup=get_reply_command_keyboard())
            return
        workout_plan = await format_workout_plan(user_id=message.from_user.id, week=week, day="среда")
        await message.answer(
            workout_plan,
            reply_markup=get_days_only_keyboard()
        )
        logger.info(f"Пользователь {message.from_user.id} запросил программу на среду, неделя {week}")
    except Exception as e:
        logger.error(f"Ошибка в wednesday_button для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте снова.", reply_markup=get_reply_command_keyboard())


# Обработчик кнопки "пятница"
@router.message(MaxLiftForm.week_selection, lambda message: message.text == "пятница")
async def friday_button(message: Message, state: FSMContext):
    """Обрабатывает нажатие кнопки 'пятница'."""
    try:
        data = await state.get_data()
        week = data.get("selected_week")
        if not week:
            await message.answer("Ошибка: неделя не выбрана. Используй кнопку 'неделя'.",
                                 reply_markup=get_reply_command_keyboard())
            return
        workout_plan = await format_workout_plan(user_id=message.from_user.id, week=week, day="пятница")
        await message.answer(
            workout_plan,
            reply_markup=get_days_only_keyboard()
        )
        logger.info(f"Пользователь {message.from_user.id} запросил программу на пятницу, неделя {week}")
    except Exception as e:
        logger.error(f"Ошибка в friday_button для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте снова.", reply_markup=get_reply_command_keyboard())


# Обработчик кнопки "Назад"
@router.message(MaxLiftForm.week_selection, lambda message: message.text == "Назад")
async def back_button(message: Message, state: FSMContext):
    """Обрабатывает нажатие кнопки 'Назад' для возврата к главному меню."""
    try:
        await state.clear()
        await message.answer(
            "Выбери действие:",
            reply_markup=get_reply_command_keyboard()
        )
        logger.info(f"Пользователь {message.from_user.id} вернулся к главному меню")
    except Exception as e:
        logger.error(f"Ошибка в back_button для пользователя {message.from_user.id}: {e}")
        await message.answer("Произошла ошибка. Попробуйте снова.", reply_markup=get_reply_command_keyboard())


# Основная функция для запуска бота
async def main():
    """Запускает бота и инициализирует базу данных."""
    try:
        init_db()
        await bot.delete_webhook(drop_pending_updates=True)
        logger.info("Бот запущен...")
        await dp.start_polling(bot)
    except Exception as e:
        logger.error(f"Ошибка при запуске бота: {e}")
        raise


if __name__ == "__main__":
    asyncio.run(main())